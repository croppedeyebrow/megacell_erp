"""레거시 엑셀(자재/구매/BOM/수주) → 신규 DB 적재.

대상 파일 (data/legacy_imports/ 기준):
  - 자재리스트 관리 양식__v0.8.xlsm  (제조사리스트/부품리스트/입출고내역/BOM 관리)
  - 자재발주현황_발주서 통합 양식.xlsm (발주현황/업체 이름_번호)
  - PDP 파트리스트 정리.xlsx         (제품별 BOM, 시트당 1개 완제품)
  - 영업_2026년 수주관리대장.xlsx     (2026년 수주 원장)

각 함수는 ImportJob 1건을 만들고, 실패한 행은 건너뛰며 상세를 details_json 에 기록합니다.
같은 소스를 다시 실행해도 안전하도록(재실행 시 중복 적재 방지) upsert 방식을 사용합니다.
"""

from __future__ import annotations

import datetime as dt
from decimal import Decimal
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.core.security import utcnow
from app.domains.imports.mapping import to_decimal
from app.domains.inventory.models import MaterialMovement
from app.domains.master_data.models import BomComponent, Material, Supplier
from app.domains.procurement.models import PurchaseOrder
from app.domains.sales.models import ImportJob, SalesOrder


class ExcelImportError(Exception):
    def __init__(self, message: str) -> None:
        self.message = message
        super().__init__(message)


# ---------------------------------------------------------------------------
# 공통 헬퍼
# ---------------------------------------------------------------------------


def _open(path: Path) -> openpyxl.Workbook:
    if not path.exists():
        raise ExcelImportError(f"파일을 찾을 수 없습니다: {path}")
    return openpyxl.load_workbook(path, data_only=True)


def _clean_str(v: Any) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def _find_header(
    ws: Worksheet, anchor_text: str, max_scan: int = 20
) -> tuple[int, dict[str, int]]:
    """anchor_text 가 등장하는 행을 헤더 행으로 보고, 1~해당 행까지의
    문자열 셀들을 모두 모아 {라벨: 열인덱스(0-base)} 맵을 만든다.

    여러 줄에 걸쳐 라벨이 흩어져 있는 시트(예: 발주현황, 입출고내역)를 처리하기 위함.
    """
    header_row = None
    for row_idx, row in enumerate(
        ws.iter_rows(min_row=1, max_row=max_scan, values_only=True), start=1
    ):
        for cell in row:
            if isinstance(cell, str) and cell.strip() == anchor_text:
                header_row = row_idx
                break
        if header_row:
            break
    if header_row is None:
        raise ExcelImportError(
            f"'{ws.title}' 시트에서 헤더 기준 '{anchor_text}' 를 찾지 못했습니다."
        )

    colmap: dict[str, int] = {}
    for row in ws.iter_rows(min_row=1, max_row=header_row, values_only=True):
        for i, v in enumerate(row):
            if isinstance(v, str) and v.strip():
                colmap[v.strip()] = i
    return header_row, colmap


def _get(row: tuple[Any, ...], colmap: dict[str, int], label: str) -> Any:
    idx = colmap.get(label)
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def _to_price(v: Any) -> Decimal | None:
    """단가/금액 셀. '#VALUE!' 같은 수식 오류 문자열은 무시."""
    if isinstance(v, str) and (v.startswith("#") or v.strip() in {"-", ""}):
        return None
    return to_decimal(v)


def _to_date_str(v: Any) -> str | None:
    if v is None:
        return None
    if isinstance(v, (dt.datetime, dt.date)):
        return v.date().isoformat() if isinstance(v, dt.datetime) else v.isoformat()
    if isinstance(v, str):
        s = v.strip()
        return s or None
    return None


def _decode_yymmdd(v: Any) -> str | None:
    """'영업_2026년 수주관리대장' 시트의 2250417 형태(=2025-04-17) 정수 날짜 디코딩."""
    if v is None:
        return None
    if isinstance(v, (dt.datetime, dt.date)):
        return _to_date_str(v)
    try:
        s = str(int(v)).strip()
    except (ValueError, TypeError):
        return _clean_str(v)
    if len(s) == 7 and s.startswith("2"):
        core = s[1:]
    elif len(s) == 6:
        core = s
    else:
        return _clean_str(v)
    yy, mm, dd = core[0:2], core[2:4], core[4:6]
    try:
        return dt.date(2000 + int(yy), int(mm), int(dd)).isoformat()
    except ValueError:
        return None


def _new_job(db: Session, job_type: str, source_path: Path, started_by: str | None) -> ImportJob:
    job = ImportJob(
        job_type=job_type,
        source_path=str(source_path),
        status="processing",
        started_by=started_by,
        started_at=utcnow(),
    )
    db.add(job)
    db.flush()
    return job


def _finish_job(
    db: Session,
    job: ImportJob,
    *,
    success: int,
    warnings: int,
    fails: int,
    message: str,
    details: dict[str, Any],
) -> ImportJob:
    job.status = "partial_success" if (fails or warnings) else "success"
    if fails and success == 0:
        job.status = "failed"
    job.success_count = success
    job.warning_count = warnings
    job.fail_count = fails
    job.message = message
    job.details_json = details
    job.ended_at = utcnow()
    db.commit()
    db.refresh(job)
    return job


# ---------------------------------------------------------------------------
# 1) 공급업체 (제조사리스트 + 업체 이름_번호)
# ---------------------------------------------------------------------------


def _upsert_supplier(
    db: Session,
    cache: dict[str, Supplier],
    *,
    name: str,
    tel: str | None,
    mobile: str | None,
    fax: str | None,
    main_items: str | None,
    note: str | None,
    source_file: str,
) -> Supplier:
    key = name.strip()
    existing = cache.get(key)
    if existing is None:
        existing = db.scalar(select(Supplier).where(Supplier.name == key))
    if existing:
        if tel and not existing.tel:
            existing.tel = tel
        if mobile and not existing.mobile:
            existing.mobile = mobile
        if fax and not existing.fax:
            existing.fax = fax
        if main_items and not existing.main_items:
            existing.main_items = main_items
        if note and not existing.note:
            existing.note = note
        cache[key] = existing
        return existing

    supplier = Supplier(
        name=key,
        tel=tel,
        mobile=mobile,
        fax=fax,
        main_items=main_items,
        note=note,
        source_file=source_file,
    )
    db.add(supplier)
    db.flush()
    cache[key] = supplier
    return supplier


def import_suppliers(
    db: Session, *, materials_path: Path, procurement_path: Path, started_by: str | None
) -> ImportJob:
    job = _new_job(db, "excel_suppliers", materials_path, started_by)
    success = 0
    fails = 0
    details: dict[str, Any] = {"errors": []}
    cache: dict[str, Supplier] = {}

    try:
        wb1 = _open(materials_path)
        ws1 = wb1["제조사리스트"]
        header_row, colmap = _find_header(ws1, "No.")
        for row in ws1.iter_rows(min_row=header_row + 1, max_row=ws1.max_row, values_only=True):
            name = _clean_str(_get(row, colmap, "제조사"))
            if not name:
                continue
            try:
                _upsert_supplier(
                    db,
                    cache,
                    name=name,
                    tel=_clean_str(_get(row, colmap, "연락처")),
                    mobile=None,
                    fax=_clean_str(_get(row, colmap, "FAX")),
                    main_items=_clean_str(_get(row, colmap, "제작 품목")),
                    note=_clean_str(_get(row, colmap, "비고")),
                    source_file=materials_path.name,
                )
                success += 1
            except Exception as exc:  # noqa: BLE001
                fails += 1
                details["errors"].append({"sheet": "제조사리스트", "error": str(exc)})

        wb2 = _open(procurement_path)
        ws2 = wb2["업체 이름_번호"]
        header_row2, colmap2 = _find_header(ws2, "업체명")
        for row in ws2.iter_rows(min_row=header_row2 + 1, max_row=ws2.max_row, values_only=True):
            name = _clean_str(_get(row, colmap2, "업체명"))
            if not name:
                continue
            try:
                _upsert_supplier(
                    db,
                    cache,
                    name=name,
                    tel=_clean_str(_get(row, colmap2, "TEL")),
                    mobile=_clean_str(_get(row, colmap2, "H.P")),
                    fax=_clean_str(_get(row, colmap2, "FAX")),
                    main_items=None,
                    note=None,
                    source_file=procurement_path.name,
                )
                success += 1
            except Exception as exc:  # noqa: BLE001
                fails += 1
                details["errors"].append({"sheet": "업체 이름_번호", "error": str(exc)})

        db.commit()
        details["supplier_count_after"] = db.query(Supplier).count()
        return _finish_job(
            db,
            job,
            success=success,
            warnings=0,
            fails=fails,
            message=f"공급업체 {success}건 처리 / 실패 {fails}",
            details=details,
        )
    except Exception as exc:
        db.rollback()
        return _finish_job(
            db, job, success=0, warnings=0, fails=1, message=str(exc), details=details
        )


# ---------------------------------------------------------------------------
# 2) 자재 마스터 (부품리스트)
# ---------------------------------------------------------------------------


def import_materials(db: Session, *, path: Path, started_by: str | None) -> ImportJob:
    job = _new_job(db, "excel_materials", path, started_by)
    success = 0
    fails = 0
    details: dict[str, Any] = {"errors": []}

    try:
        wb = _open(path)
        ws = wb["부품리스트"]
        header_row, colmap = _find_header(ws, "자재품목")

        supplier_cache: dict[str, str | None] = {}

        for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row, values_only=True):
            category = _clean_str(_get(row, colmap, "자재품목"))
            name = _clean_str(_get(row, colmap, "자재품명"))
            if not category or not name:
                continue
            try:
                manufacturer_name = _clean_str(_get(row, colmap, "제조사"))
                supplier_id = None
                if manufacturer_name:
                    if manufacturer_name not in supplier_cache:
                        sup = db.scalar(select(Supplier).where(Supplier.name == manufacturer_name))
                        supplier_cache[manufacturer_name] = sup.id if sup else None
                    supplier_id = supplier_cache[manufacturer_name]

                existing = db.scalar(
                    select(Material).where(Material.category == category, Material.name == name)
                )
                payload = dict(
                    category=category,
                    name=name,
                    manufacturer_name=manufacturer_name,
                    supplier_id=supplier_id,
                    unit=_clean_str(_get(row, colmap, "단위")),
                    unit_price=_to_price(_get(row, colmap, "단가")),
                    usage_note=_clean_str(_get(row, colmap, "자재용도")),
                    source_file=path.name,
                )
                if existing:
                    for k, v in payload.items():
                        setattr(existing, k, v)
                else:
                    db.add(Material(**payload))
                success += 1
            except Exception as exc:  # noqa: BLE001
                fails += 1
                details["errors"].append({"category": category, "name": name, "error": str(exc)})

        db.commit()
        return _finish_job(
            db,
            job,
            success=success,
            warnings=0,
            fails=fails,
            message=f"자재 {success}건 처리 / 실패 {fails}",
            details=details,
        )
    except Exception as exc:
        db.rollback()
        return _finish_job(
            db, job, success=0, warnings=0, fails=1, message=str(exc), details=details
        )


# ---------------------------------------------------------------------------
# 3) BOM (자재리스트!BOM 관리 + PDP 파트리스트 정리 시트별)
# ---------------------------------------------------------------------------


def _material_id_for(db: Session, cache: dict[tuple[str, str], str | None], category: str | None, name: str | None) -> str | None:
    if not name:
        return None
    key = (category or "", name)
    if key not in cache:
        q = select(Material).where(Material.name == name)
        if category:
            q = q.where(Material.category == category)
        mat = db.scalar(q)
        cache[key] = mat.id if mat else None
    return cache[key]


def import_bom(
    db: Session, *, materials_path: Path, pdp_path: Path | None, started_by: str | None
) -> ImportJob:
    job = _new_job(db, "excel_bom", materials_path, started_by)
    success = 0
    fails = 0
    details: dict[str, Any] = {"errors": [], "sources": []}
    material_cache: dict[tuple[str, str], str | None] = {}

    try:
        # 기존에 이 두 소스에서 적재된 BOM은 다시 실행 시 중복되지 않도록 먼저 제거
        db.query(BomComponent).filter(
            BomComponent.source_file.in_([materials_path.name, pdp_path.name if pdp_path else "__none__"])
        ).delete(synchronize_session=False)

        wb = _open(materials_path)
        ws = wb["BOM 관리"]
        header_row, colmap = _find_header(ws, "완제품명")
        details["sources"].append(f"{materials_path.name}!BOM 관리")
        for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row, values_only=True):
            product_name = _clean_str(_get(row, colmap, "완제품명"))
            material_name = _clean_str(_get(row, colmap, "자재품명"))
            if not product_name or not material_name:
                continue
            try:
                material_category = _clean_str(_get(row, colmap, "자재품목"))
                db.add(
                    BomComponent(
                        product_category=_clean_str(_get(row, colmap, "제품 분류")),
                        product_name=product_name,
                        material_category=material_category,
                        material_name=material_name,
                        material_id=_material_id_for(db, material_cache, material_category, material_name),
                        manufacturer_name=_clean_str(_get(row, colmap, "제조사")),
                        unit=_clean_str(_get(row, colmap, "단위")),
                        quantity=to_decimal(_get(row, colmap, "Q'ty")),
                        unit_price=_to_price(_get(row, colmap, "단가")),
                        usage_note=_clean_str(_get(row, colmap, "자재용도")),
                        source_file=materials_path.name,
                        source_sheet="BOM 관리",
                    )
                )
                success += 1
            except Exception as exc:  # noqa: BLE001
                fails += 1
                details["errors"].append({"product": product_name, "material": material_name, "error": str(exc)})

        if pdp_path is not None:
            wb2 = _open(pdp_path)
            for sheet_name in wb2.sheetnames:
                ws2 = wb2[sheet_name]
                try:
                    header_row2, colmap2 = _find_header(ws2, "No")
                except ExcelImportError as exc:
                    details["errors"].append({"sheet": sheet_name, "error": str(exc)})
                    continue
                product_category = sheet_name.split("_")[0].strip() if "_" in sheet_name else sheet_name
                details["sources"].append(f"{pdp_path.name}!{sheet_name}")
                for row in ws2.iter_rows(
                    min_row=header_row2 + 1, max_row=ws2.max_row, values_only=True
                ):
                    material_category = _clean_str(_get(row, colmap2, "Name. Description(품목)"))
                    material_name = _clean_str(_get(row, colmap2, "Specification(품명)"))
                    if not material_category and not material_name:
                        continue
                    try:
                        qty = _get(row, colmap2, "필요수량")
                        if qty is None:
                            qty = _get(row, colmap2, "Q'ty")
                        db.add(
                            BomComponent(
                                product_category=product_category,
                                product_name=sheet_name,
                                material_category=material_category,
                                material_name=material_name or material_category or "미상",
                                material_id=_material_id_for(
                                    db, material_cache, material_category, material_name
                                ),
                                manufacturer_name=_clean_str(_get(row, colmap2, "납품업체")),
                                unit=_clean_str(_get(row, colmap2, "Unit(단위)")),
                                quantity=to_decimal(qty),
                                unit_price=_to_price(_get(row, colmap2, "단가")),
                                usage_note=_clean_str(_get(row, colmap2, "Remark")),
                                source_file=pdp_path.name,
                                source_sheet=sheet_name,
                            )
                        )
                        success += 1
                    except Exception as exc:  # noqa: BLE001
                        fails += 1
                        details["errors"].append(
                            {"sheet": sheet_name, "material": material_name, "error": str(exc)}
                        )

        db.commit()
        return _finish_job(
            db,
            job,
            success=success,
            warnings=0,
            fails=fails,
            message=f"BOM 구성 {success}건 처리 / 실패 {fails}",
            details=details,
        )
    except Exception as exc:
        db.rollback()
        return _finish_job(
            db, job, success=0, warnings=0, fails=1, message=str(exc), details=details
        )


# ---------------------------------------------------------------------------
# 4) 자재 발주 (자재발주현황!발주현황)
# ---------------------------------------------------------------------------


def import_purchase_orders(db: Session, *, path: Path, started_by: str | None) -> ImportJob:
    job = _new_job(db, "excel_purchase_orders", path, started_by)
    success = 0
    fails = 0
    warnings = 0
    details: dict[str, Any] = {"errors": []}

    try:
        wb = _open(path)
        ws = wb["발주현황"]
        header_row, colmap = _find_header(ws, "업체명")

        for idx, row in enumerate(
            ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row, values_only=True), start=header_row + 1
        ):
            supplier_name = _clean_str(_get(row, colmap, "업체명"))
            if not supplier_name:
                continue
            try:
                source_key = _clean_str(_get(row, colmap, "발주연동키"))
                if not source_key:
                    source_key = f"{path.name}:발주현황:{idx}"
                    warnings += 1

                supplier = db.scalar(select(Supplier).where(Supplier.name == supplier_name))

                existing = db.scalar(
                    select(PurchaseOrder).where(PurchaseOrder.source_key == source_key)
                )
                payload = dict(
                    seq_no=int(_get(row, colmap, "No.")) if _get(row, colmap, "No.") is not None else None,
                    order_no=_clean_str(_get(row, colmap, "발주번호")),
                    supplier_name=supplier_name,
                    supplier_id=supplier.id if supplier else None,
                    order_date=_to_date_str(_get(row, colmap, "발주일")),
                    due_date=_to_date_str(_get(row, colmap, "입고(요청)일")),
                    material_category=_clean_str(_get(row, colmap, "품목")),
                    material_name=_clean_str(_get(row, colmap, "품명")),
                    quantity=to_decimal(_get(row, colmap, "개수")),
                    unit_price=_to_price(_get(row, colmap, "단가")),
                    total_amount=_to_price(_get(row, colmap, "총금액")),
                    purpose=_clean_str(_get(row, colmap, "발주목적")),
                    status=_clean_str(_get(row, colmap, "진행 상태")),
                    memo=_clean_str(_get(row, colmap, "비고")),
                    source_key=source_key,
                    source_file=path.name,
                    source_row=idx,
                )
                if existing:
                    for k, v in payload.items():
                        setattr(existing, k, v)
                else:
                    db.add(PurchaseOrder(**payload))
                success += 1
            except Exception as exc:  # noqa: BLE001
                fails += 1
                details["errors"].append({"row": idx, "supplier": supplier_name, "error": str(exc)})

        db.commit()
        return _finish_job(
            db,
            job,
            success=success,
            warnings=warnings,
            fails=fails,
            message=f"발주 {success}건 처리 / 발주연동키 없음 {warnings} / 실패 {fails}",
            details=details,
        )
    except Exception as exc:
        db.rollback()
        return _finish_job(
            db, job, success=0, warnings=0, fails=1, message=str(exc), details=details
        )


# ---------------------------------------------------------------------------
# 5) 자재 입출고 내역 (자재리스트!입출고내역)
# ---------------------------------------------------------------------------


def import_material_movements(db: Session, *, path: Path, started_by: str | None) -> ImportJob:
    job = _new_job(db, "excel_material_movements", path, started_by)
    success = 0
    fails = 0
    details: dict[str, Any] = {"errors": []}

    try:
        wb = _open(path)
        ws = wb["입출고내역"]
        header_row, colmap = _find_header(ws, "원본파일")

        material_cache: dict[tuple[str, str], str | None] = {}

        for idx, row in enumerate(
            ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row, values_only=True), start=header_row + 1
        ):
            material_name = _clean_str(_get(row, colmap, "자재품명"))
            if not material_name:
                continue
            try:
                material_category = _clean_str(_get(row, colmap, "자재품목"))
                existing = db.scalar(
                    select(MaterialMovement).where(
                        MaterialMovement.source_file == path.name,
                        MaterialMovement.source_sheet == "입출고내역",
                        MaterialMovement.source_row == idx,
                    )
                )
                payload = dict(
                    moved_at=_to_date_str(_get(row, colmap, "입출고일")),
                    order_no=_clean_str(_get(row, colmap, "발주번호")),
                    material_category=material_category,
                    material_name=material_name,
                    material_id=_material_id_for(db, material_cache, material_category, material_name),
                    purpose=_clean_str(_get(row, colmap, "발주목적/출고목적")),
                    manufacturer_name=_clean_str(_get(row, colmap, "제조사")),
                    unit=_clean_str(_get(row, colmap, "단위")),
                    in_qty=to_decimal(_get(row, colmap, "입고수량")),
                    out_qty=to_decimal(_get(row, colmap, "출고수량")),
                    unit_price=_to_price(_get(row, colmap, "단가")),
                    amount=_to_price(_get(row, colmap, "금액")),
                    purchase_order_source_key=_clean_str(_get(row, colmap, "발주연동키")),
                    source_file=path.name,
                    source_sheet="입출고내역",
                    source_row=idx,
                )
                if existing:
                    for k, v in payload.items():
                        setattr(existing, k, v)
                else:
                    db.add(MaterialMovement(**payload))
                success += 1
            except Exception as exc:  # noqa: BLE001
                fails += 1
                details["errors"].append({"row": idx, "material": material_name, "error": str(exc)})

        db.commit()
        return _finish_job(
            db,
            job,
            success=success,
            warnings=0,
            fails=fails,
            message=f"자재 입출고 {success}건 처리 / 실패 {fails}",
            details=details,
        )
    except Exception as exc:
        db.rollback()
        return _finish_job(
            db, job, success=0, warnings=0, fails=1, message=str(exc), details=details
        )


# ---------------------------------------------------------------------------
# 6) 2026년 수주관리대장 → sales_orders
# ---------------------------------------------------------------------------

_ORDER_CATEGORY_LABELS = {
    1: "정류기",
    2: "UPS",
    3: "인버터",
    4: "축전지",
    5: "기타",
    6: "외주",
}

_INVOICE_STATUS_LABELS = {
    "M": "계산서 발행",
    "M-1": "계산서 미발행",
    "COU": "쿠팡",
    "X": "계산서 발행안함",
}


def import_sales_orders_2026(db: Session, *, path: Path, started_by: str | None) -> ImportJob:
    job = _new_job(db, "excel_sales_orders_2026", path, started_by)
    success = 0
    fails = 0
    warnings = 0
    details: dict[str, Any] = {"errors": []}

    try:
        wb = _open(path)
        ws = wb["2026년수주"]
        header_row, colmap = _find_header(ws, "제조지시서")

        for idx, row in enumerate(
            ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row, values_only=True), start=header_row + 1
        ):
            product_name = _clean_str(_get(row, colmap, "품명 및 규격"))
            customer_name = _clean_str(_get(row, colmap, "발주처"))
            if not product_name and not customer_name:
                continue
            try:
                order_no_raw = _clean_str(_get(row, colmap, "제조지시서"))
                source_key = f"excel2026:{path.name}:{idx}"

                order_no = order_no_raw or f"EXCEL2026-{idx}"
                clash = db.scalar(select(SalesOrder).where(SalesOrder.order_no == order_no))
                if clash and clash.source_key != source_key:
                    order_no = f"{order_no}-{idx}"
                    warnings += 1

                # 주의: 이 시트는 헤더에 '구분' 라벨이 두 번 나온다(수주 구분 1~6 / 담당자 옆 잉여 라벨).
                # colmap 은 마지막 라벨 위치로 덮어써지므로, 실제 수주구분 코드는 고정 위치(B열, index 1)에서 읽는다.
                cat_code = row[1] if len(row) > 1 else None
                category = _ORDER_CATEGORY_LABELS.get(int(cat_code)) if isinstance(cat_code, (int, float)) else None

                order_amount = _to_price(_get(row, colmap, "수주금액"))
                vat_incl = _to_price(_get(row, colmap, "V.A.T 포함"))
                vat_amount = (vat_incl - order_amount) if (order_amount is not None and vat_incl is not None) else None

                invoice_code = _clean_str(_get(row, colmap, "계산서"))
                invoice_status = _INVOICE_STATUS_LABELS.get(invoice_code or "", invoice_code)

                delivery_site = _clean_str(_get(row, colmap, "납품처"))
                contact = _clean_str(_get(row, colmap, "발주처담당"))
                memo_parts = [p for p in [f"납품처: {delivery_site}" if delivery_site else None,
                                           f"발주처담당: {contact}" if contact else None,
                                           _clean_str(_get(row, colmap, "비고"))] if p]
                memo = " / ".join(memo_parts) or None

                existing = db.scalar(select(SalesOrder).where(SalesOrder.source_key == source_key))
                payload = dict(
                    order_no=order_no,
                    source_key=source_key,
                    order_date=_decode_yymmdd(_get(row, colmap, "수주일자")),
                    category=category,
                    customer_name=customer_name or "",
                    vendor_name=None,
                    product_name=product_name or "",
                    unit=_clean_str(_get(row, colmap, "단위")),
                    quantity=to_decimal(_get(row, colmap, "수량")),
                    unit_price=_to_price(_get(row, colmap, "단가")),
                    order_amount=order_amount,
                    vat_amount=vat_amount,
                    invoice_status=invoice_status,
                    due_date=_decode_yymmdd(_get(row, colmap, "출고요청일")),
                    shipped_date=_decode_yymmdd(_get(row, colmap, "출고일")),
                    status="completed" if _get(row, colmap, "출고일") else "confirmed",
                    memo=memo,
                    # 담당자명은 헤더에 라벨이 없는 시트 맨 끝 열에 있다(원본 관측 기준).
                    owner_name=_clean_str(row[-1]) if row else None,
                )

                if existing:
                    for k, v in payload.items():
                        setattr(existing, k, v)
                    existing.version = int(existing.version or 1) + 1
                else:
                    db.add(SalesOrder(**payload))
                success += 1
            except Exception as exc:  # noqa: BLE001
                fails += 1
                details["errors"].append({"row": idx, "product": product_name, "error": str(exc)})

        db.commit()
        return _finish_job(
            db,
            job,
            success=success,
            warnings=warnings,
            fails=fails,
            message=f"2026년 수주 {success}건 처리 / 주문번호 충돌 {warnings} / 실패 {fails}",
            details=details,
        )
    except Exception as exc:
        db.rollback()
        return _finish_job(
            db, job, success=0, warnings=0, fails=1, message=str(exc), details=details
        )


# ---------------------------------------------------------------------------
# 오케스트레이션
# ---------------------------------------------------------------------------


def run_priority_imports(db: Session, *, data_dir: Path, started_by: str | None = None) -> list[ImportJob]:
    materials_path = data_dir / "자재리스트 관리 양식__v0.8.xlsm"
    procurement_path = data_dir / "자재발주현황_발주서 통합 양식.xlsm"
    pdp_path = data_dir / "PDP 파트리스트 정리.xlsx"
    sales_path = data_dir / "영업_2026년 수주관리대장.xlsx"

    jobs: list[ImportJob] = []
    jobs.append(
        import_suppliers(
            db, materials_path=materials_path, procurement_path=procurement_path, started_by=started_by
        )
    )
    jobs.append(import_materials(db, path=materials_path, started_by=started_by))
    jobs.append(
        import_bom(
            db,
            materials_path=materials_path,
            pdp_path=pdp_path if pdp_path.exists() else None,
            started_by=started_by,
        )
    )
    jobs.append(import_purchase_orders(db, path=procurement_path, started_by=started_by))
    jobs.append(import_material_movements(db, path=materials_path, started_by=started_by))
    if sales_path.exists():
        jobs.append(import_sales_orders_2026(db, path=sales_path, started_by=started_by))
    return jobs
