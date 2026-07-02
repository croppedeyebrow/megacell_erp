from __future__ import annotations

import io
import re
from datetime import date

import pandas as pd

from config import BASE_DIR, TEMPLATES_DIR


def parse_int(value: object) -> int:
    text = "" if pd.isna(value) else str(value)
    text = text.replace(",", "").replace("원", "").strip()
    number = pd.to_numeric(text, errors="coerce")
    return 0 if pd.isna(number) else int(number)

def first_match(pattern: str, text: str, default: str = "", flags: int = 0) -> str:
    match = re.search(pattern, text, flags)
    return match.group(1).strip() if match else default

def format_date_parts(year: str, month: str, day: str) -> str:
    return f"{int(year):04}.{int(month):02}.{int(day):02}"

def default_order_no() -> str:
    today = date.today()
    return f"2{today:%y%m%d}-01"

def normalize_delivery_date(raw: str, quote_date: str) -> str:
    text = raw.strip()
    year = quote_date[:4] if re.match(r"\d{4}\.", quote_date) else str(date.today().year)
    match = re.match(r"(\d{1,2})월\s*(\d{1,2})일(.*)", text)
    if not match:
        return text
    return f"{year}.{int(match.group(1)):02}.{int(match.group(2)):02}{match.group(3)}"

def build_order_item_name(lines: list[str]) -> str:
    if not lines:
        return ""

    joined = " ".join(line.strip() for line in lines if line.strip())
    product_line = lines[0].strip()
    model_match = re.search(r"/\s*([A-Za-z0-9-]+)", product_line)
    spec_match = re.search(r"규격:\s*([^\s(]+)", joined)
    cell_match = re.search(r"\*+\s*(\d+\s*Cell)", joined, re.IGNORECASE)
    rack_text = "일체형 랙 구성" if "일체형 랙 구성" in joined else ""
    height_match = re.search(r"\*(\d{3,4})mm", joined)

    if product_line.startswith("축전지") and model_match and spec_match:
        cell_text = f" * {cell_match.group(1).replace(' ', '')}" if cell_match else ""
        rack_suffix = f" + {rack_text}" if rack_text else ""
        height_suffix = f" {height_match.group(1)}mm" if height_match else ""
        return f"축전지 / {spec_match.group(1)}({model_match.group(1)}{cell_text}){rack_suffix}{height_suffix}"

    return joined

def extract_quote_items(lines: list[str]) -> list[dict[str, object]]:
    items: list[dict[str, object]] = []
    buffer: list[str] = []
    in_table = False

    for raw_line in lines:
        line = raw_line.strip()
        if not line:
            continue
        if "Category Product Name" in line:
            in_table = True
            buffer = []
            continue
        if not in_table:
            continue
        if line.startswith("결 재") or line.startswith("납 ") or line.startswith("제 품") or line.startswith("상 호"):
            break
        if "합계(V.A.T" in line or line == "-" or re.fullmatch(r"[-\s]+", line):
            continue

        inline_match = re.match(r"(.+?)\s+(\d+)\s+([\d,]+)\s+([\d,]+)\s*$", line)
        if inline_match:
            desc = inline_match.group(1).strip()
            qty = parse_int(inline_match.group(2))
            unit_price = parse_int(inline_match.group(3))
            total = parse_int(inline_match.group(4))
            if desc:
                item_name = desc
            else:
                item_name = build_order_item_name(buffer)
            if item_name:
                items.append(
                    {
                        "품명규격": item_name,
                        "수량": qty,
                        "단가": unit_price,
                        "금액": total,
                    }
                )
            buffer = []
            continue

        amount_match = re.match(r"(\d+)\s+([\d,]+)\s+([\d,]+)\s*$", line)
        if amount_match:
            item_name = build_order_item_name(buffer)
            if item_name:
                items.append(
                    {
                        "품명규격": item_name,
                        "수량": parse_int(amount_match.group(1)),
                        "단가": parse_int(amount_match.group(2)),
                        "금액": parse_int(amount_match.group(3)),
                    }
                )
            buffer = []
            continue

        if not re.fullmatch(r"[\d,]+", line):
            buffer.append(line)

    return items

def extract_pdf_text(pdf_bytes: bytes) -> tuple[str, list[str]]:
    try:
        from pypdf import PdfReader
    except ImportError as exc:
        raise RuntimeError("PDF 처리 라이브러리(pypdf)가 없습니다. requirements.txt 설치 또는 앱 폴더의 pypdf 포함이 필요합니다.") from exc

    reader = PdfReader(io.BytesIO(pdf_bytes))
    text = "\n".join(page.extract_text() or "" for page in reader.pages)
    lines = [line.strip() for line in text.splitlines() if line.strip()]
    return text, lines

def extract_quote_pdf(pdf_bytes: bytes) -> dict[str, object]:
    text, lines = extract_pdf_text(pdf_bytes)

    quote_no = first_match(r"NO:\s*(MGC-\s*\d+)", text).replace(" ", "")
    date_match = re.search(r"(\d{4})\s*년\s*(\d{1,2})\s*월\s*(\d{1,2})\s*일", text)
    quote_date = format_date_parts(*date_match.groups()) if date_match else ""

    recipient = first_match(r"\n\s*([^\n]+?)\s+귀중", text)
    customer = ""
    contact = ""
    if "/" in recipient:
        customer, contact = [part.strip() for part in recipient.split("/", 1)]
    else:
        customer = recipient.strip()
    contact = contact.replace("님", "").strip()

    site = first_match(r"\n\s*\(([^()\n]+)\)\s*\n\s*아래", text)
    delivery_raw = first_match(r"납\s*기\s*:\s*([^\n]+)", text)
    delivery_place_raw = first_match(r"납\s*품\s*장\s*소\s*:\s*([^\n]+)", text)
    delivery_place = delivery_place_raw
    receiver = ""
    if "담당자:" in delivery_place_raw:
        delivery_place, receiver = [part.strip() for part in delivery_place_raw.split("담당자:", 1)]
        delivery_place = delivery_place.rstrip(",").strip()

    product_config = first_match(r"제\s*품\s*구\s*성\s*:\s*([^\n]+)", text)
    total_amount = parse_int(first_match(r"합계\(V\.A\.T별도\)\s*([\d,]+)", text))
    items = extract_quote_items(lines)
    if "폐전지 회수" in product_config:
        for item in items:
            name = str(item.get("품명규격", "")).strip()
            if name.startswith("설치공사비") and "폐전지" not in name:
                detail = name.replace("설치공사비", "", 1).strip()
                item["품명규격"] = f"설치공사비({detail}+ 폐전지 회수)" if detail else "설치공사비(폐전지 회수)"
    if not total_amount:
        total_amount = sum(parse_int(item.get("금액")) for item in items)

    return {
        "견적번호": quote_no,
        "견적일자": quote_date,
        "발주처": customer,
        "담당자": contact,
        "현장명": site,
        "발행일자": date.today().strftime("%Y.%m.%d"),
        "납품일자": normalize_delivery_date(delivery_raw, quote_date),
        "주문번호": default_order_no(),
        "납품장소": delivery_place,
        "현지인수자": receiver,
        "제품정보": product_config,
        "합계금액": total_amount,
        "품목": items,
    }

def find_purchase_company(text: str) -> str:
    matches = re.findall(r"\n\s*((?:㈜|\(주\)|주식회사)\s*[^\n]+)\s*\n\s*\d{3}-\d{2}-\d{5}", text)
    for value in matches:
        company = re.sub(r"\s+", " ", value).strip()
        if "메가셀" not in company:
            return company
    return ""

def find_purchase_contact(text: str, customer: str) -> str:
    if customer:
        pattern = re.escape(customer) + r".{0,220}?\n\s*([^\n]*(?:팀|부|과|실)[^\n]*(?:부장|팀장|과장|차장|대리|사원)?)\s*\n\s*([\d,-]+[^\n]*)"
        match = re.search(pattern, text, re.S)
        if match:
            return f"{match.group(1).strip()} / {match.group(2).strip()}"
    match = re.search(r"\n\s*([^\n]*(?:정순열|정재헌)[^\n]*)\s*\n\s*([\d,-]+[^\n]*)", text)
    if match:
        return f"{match.group(1).strip()} / {match.group(2).strip()}"
    return ""

def find_purchase_contract_name(text: str, lines: list[str]) -> str:
    inline = first_match(r"계약명\s+([^\n]+)", text)
    if inline and not inline.startswith(("계약일자", "제품명")) and not re.match(r"\d{4}[.-]\d{2}", inline):
        return inline

    for idx, line in enumerate(lines):
        if line.startswith("부가가치세") and idx > 0:
            candidate = lines[idx - 1].strip()
            if candidate and not re.search(r"^\d{4}[.-]\d{2}", candidate):
                return candidate

    for line in lines:
        if any(token in line for token in ["시설공급", "시스템", "공사", "납품", "구매"]) and not line.startswith(("제품명", "특기사항")):
            if not re.search(r"\d{1,3},\d{3}", line):
                return line.strip()
    return ""

def find_purchase_condition(label: str, text: str, lines: list[str]) -> str:
    invalid = {"계약일자(기간)", "계약명", "제품명", "납품장소", "대금지급조건", "주식회사 메가셀"}
    if label == "하자보증기간":
        for idx, line in enumerate(lines):
            if line == label:
                for candidate in reversed(lines[max(0, idx - 4) : idx]):
                    candidate = candidate.strip()
                    if re.fullmatch(r"\d+\s*년", candidate):
                        return candidate
                if idx > 0 and lines[idx - 1].strip() not in invalid:
                    return lines[idx - 1].strip()

    inline = first_match(label + r"\s+([^\n]+)", text)
    if inline and inline not in invalid and not inline.startswith("대금지급조건"):
        return inline
    for idx, line in enumerate(lines):
        if line == label:
            for candidate in lines[idx + 1 : idx + 6]:
                candidate = candidate.strip()
                if candidate and candidate not in invalid and not re.fullmatch(r"\d{3}-\d{2}-\d{5}", candidate):
                    if label in {"납품장소", "입고요청일"} and "협의" not in candidate:
                        continue
                    return candidate
    match = re.search(r"(지엔텔\s+[^\n]+협의)", text)
    return match.group(1).strip() if match else ""

def extract_purchase_order_items(text: str, lines: list[str]) -> list[dict[str, object]]:
    product_name = ""
    for idx, line in enumerate(lines):
        if re.search(r"\s/\s*[A-Za-z0-9-]+", line) and "@" not in line and "mail" not in line.lower():
            product_name = line.strip()
            if idx + 1 < len(lines) and lines[idx + 1].strip() and not lines[idx + 1].startswith("□"):
                product_name = f"{product_name} {lines[idx + 1].strip()}"
            break

    detail_lines: list[str] = []
    collecting_detail = False
    for line in lines:
        stripped = line.strip()
        if stripped.startswith("□"):
            collecting_detail = True
            detail_lines.append(stripped)
            continue
        if collecting_detail:
            if re.search(r"/\s*[A-Za-z0-9-]+", stripped) or stripped in {"합계금액( DC포함 )", "부가가치세(VAT)", "총합계금액(VAT포함)"}:
                collecting_detail = False
                continue
            if stripped and not re.match(r"^\d+\s+\d+\s+", stripped):
                detail_lines.append(stripped)
    product_detail = "\n".join(detail_lines)

    row_match = re.search(r"\n\s*1\s+(\d+)\s+([A-Za-z가-힣]+)\s+([\d,]+)\s+([\d,]+)", text)
    if not row_match:
        return []

    return [
        {
            "품명규격": product_name or "발주 품목",
            "수량": parse_int(row_match.group(1)),
            "단가": parse_int(row_match.group(3)),
            "금액": parse_int(row_match.group(4)),
            "제품설명": product_detail,
        }
    ]

def extract_purchase_order_pdf(pdf_bytes: bytes) -> dict[str, object]:
    text, lines = extract_pdf_text(pdf_bytes)

    order_date_raw = first_match(r"발주일자[^\d]*(\d{4}[-.]\d{2}[-.]\d{2})", text)
    if not order_date_raw:
        order_date_raw = first_match(r"(\d{4}[-.]\d{2}[-.]\d{2})", text)
    order_date = order_date_raw.replace("-", ".")
    customer = find_purchase_company(text)
    contact = find_purchase_contact(text, customer)
    contract_name = find_purchase_contract_name(text, lines)
    delivery_request = find_purchase_condition("입고요청일", text, lines)
    delivery_place = find_purchase_condition("납품장소", text, lines)
    warranty = find_purchase_condition("하자보증기간", text, lines)
    payment = find_purchase_condition("대금지급조건", text, lines)
    items = extract_purchase_order_items(text, lines)

    product_info_parts = []
    if items and items[0].get("제품설명"):
        product_info_parts.append(str(items[0].get("제품설명", "")).strip())
    if contract_name:
        product_info_parts.append(f"계약명: {contract_name}")
    if warranty:
        product_info_parts.append(f"하자보증기간: {warranty}")
    if payment:
        product_info_parts.append(f"대금지급조건: {payment}")

    return {
        "문서유형": "발주서",
        "견적번호": "",
        "견적일자": order_date,
        "발주처": customer,
        "담당자": contact,
        "현장명": contract_name,
        "발행일자": order_date or date.today().strftime("%Y.%m.%d"),
        "납품일자": delivery_request,
        "주문번호": default_order_no(),
        "납품장소": delivery_place,
        "현지인수자": contact,
        "제품정보": "\n".join(part for part in product_info_parts if part),
        "합계금액": sum(parse_int(item.get("금액")) for item in items),
        "품목": items,
    }

def extract_document_pdf(pdf_bytes: bytes) -> dict[str, object]:
    text, _lines = extract_pdf_text(pdf_bytes)
    if "발주서(Purchase Order)" in text or "발주내역" in text:
        return extract_purchase_order_pdf(pdf_bytes)
    data = extract_quote_pdf(pdf_bytes)
    data["문서유형"] = "견적서"
    return data

def safe_filename(text: str) -> str:
    cleaned = re.sub(r'[\\/:*?"<>|]+', " ", text or "")
    cleaned = re.sub(r"\s+", " ", cleaned).strip()
    return cleaned or "주문서"

def build_order_workbook_bytes(data: dict[str, object], items: pd.DataFrame) -> bytes:
    import openpyxl

    template_candidates = [
        TEMPLATES_DIR / "주문서,생산지시서 양식 예시파일.xlsm",
        BASE_DIR / "주문서,생산지시서 양식 예시파일.xlsm",
    ]
    template_path = next((path for path in template_candidates if path.exists()), template_candidates[0])
    if not template_path.exists():
        raise FileNotFoundError(f"주문서 양식 파일을 찾을 수 없습니다: {template_path.name}")

    wb = openpyxl.load_workbook(template_path, keep_vba=True)
    ws = wb["주문서"]

    ws["C3"] = str(data.get("발주처", "")).strip()
    ws["C4"] = str(data.get("담당자", "")).strip()
    ws["E3"] = str(data.get("발행일자", "")).strip()
    ws["E4"] = str(data.get("납품일자", "")).strip()
    ws["C5"] = str(data.get("현장명", "")).strip()
    ws["E5"] = str(data.get("주문번호", "")).strip()
    ws["C22"] = str(data.get("제품정보", "")).strip()
    ws["C26"] = str(data.get("납품장소", "")).strip()
    ws["C28"] = str(data.get("현지인수자", "")).strip()

    for row in range(8, 20):
        ws[f"C{row}"] = None
        ws[f"D{row}"] = None
        ws[f"E{row}"] = None

    clean_items = items.copy()
    clean_items = clean_items.dropna(how="all")
    for idx, (_, item) in enumerate(clean_items.head(12).iterrows(), start=8):
        ws[f"C{idx}"] = str(item.get("품명규격", "")).strip()
        ws[f"D{idx}"] = parse_int(item.get("수량", 0))
        ws[f"E{idx}"] = parse_int(item.get("단가", 0))

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
