from __future__ import annotations

import calendar
import io
import sqlite3
import subprocess
import sys
from datetime import date
from pathlib import Path
import re
from typing import Iterable

import pandas as pd
import streamlit as st


BASE_DIR = Path(__file__).resolve().parent
DB_PATH = BASE_DIR / "megacell.db"

st.set_page_config(
    page_title="메가셀 ERP",
    page_icon="M",
    layout="wide",
    initial_sidebar_state="expanded",
)


def quote_name(name: str) -> str:
    return '"' + name.replace('"', '""') + '"'


def db_mtime() -> float:
    return DB_PATH.stat().st_mtime if DB_PATH.exists() else 0


@st.cache_data(show_spinner=False)
def list_tables(mtime: float) -> list[str]:
    if not DB_PATH.exists():
        return []
    with sqlite3.connect(DB_PATH) as conn:
        rows = conn.execute(
            "SELECT name FROM sqlite_master WHERE type='table' ORDER BY name"
        ).fetchall()
    return [row[0] for row in rows if not row[0].startswith("sqlite_")]


@st.cache_data(show_spinner=False)
def load_table(table_name: str, mtime: float) -> pd.DataFrame:
    if not DB_PATH.exists() or table_name not in list_tables(mtime):
        return pd.DataFrame()
    with sqlite3.connect(DB_PATH) as conn:
        return pd.read_sql(f"SELECT * FROM {quote_name(table_name)}", conn)


def get_table(name: str) -> pd.DataFrame:
    return load_table(name, db_mtime())


def normalize_numeric(series: pd.Series) -> pd.Series:
    cleaned = (
        series.astype(str)
        .str.replace(",", "", regex=False)
        .str.replace("원", "", regex=False)
        .str.strip()
    )
    return pd.to_numeric(cleaned, errors="coerce").fillna(0)


def prepare_numeric(df: pd.DataFrame, columns: Iterable[str]) -> pd.DataFrame:
    df = df.copy()
    for col in columns:
        if col in df.columns:
            df[col] = normalize_numeric(df[col])
    return df


def clean_display(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df
    keep = [
        col
        for col in df.columns
        if not str(col).startswith("Unnamed") and not str(col).startswith("컬럼")
    ]
    return df[keep].dropna(how="all").reset_index(drop=True)


def text_filter(
    df: pd.DataFrame,
    keyword: str,
    columns: Iterable[str] | None = None,
) -> pd.DataFrame:
    if df.empty or not keyword:
        return df
    target_cols = [col for col in (columns or df.columns) if col in df.columns]
    if not target_cols:
        target_cols = list(df.columns)
    mask = (
        df[target_cols]
        .astype(str)
        .apply(lambda col: col.str.contains(keyword, case=False, na=False))
        .any(axis=1)
    )
    return df[mask]


def safe_unique(df: pd.DataFrame, column: str) -> list[str]:
    if column not in df.columns:
        return []
    values = df[column].dropna().astype(str).str.strip()
    return sorted(value for value in values.unique() if value and value != "nan")


def table_view(
    df: pd.DataFrame,
    preferred_cols: Iterable[str] | None = None,
    height: int = 560,
) -> None:
    df = clean_display(df)
    cols = [col for col in (preferred_cols or df.columns) if col in df.columns]
    if not cols:
        cols = list(df.columns)
    st.dataframe(df[cols], width="stretch", height=height, hide_index=True)


def download_csv(df: pd.DataFrame, filename: str) -> None:
    st.download_button(
        "CSV 내려받기",
        data=df.to_csv(index=False).encode("utf-8-sig"),
        file_name=filename,
        mime="text/csv",
        width="stretch",
    )


def metric_sum(df: pd.DataFrame, column: str, suffix: str = "") -> str:
    if column not in df.columns:
        return "-"
    return f"{normalize_numeric(df[column]).sum():,.0f}{suffix}"


def parse_int(value: object) -> int:
    text = "" if pd.isna(value) else str(value)
    text = text.replace(",", "").replace("원", "").strip()
    number = pd.to_numeric(text, errors="coerce")
    return 0 if pd.isna(number) else int(number)


def first_match(pattern: str, text: str, default: str = "", flags: int = 0) -> str:
    match = re.search(pattern, text, flags)
    return match.group(1).strip() if match else default


def format_date_parts(year: str, month: str, day: str) -> str:
    return f"{int(year):04}.{int(month):02}.{int(day):02}"


def default_order_no() -> str:
    today = date.today()
    return f"2{today:%y%m%d}-01"


def normalize_delivery_date(raw: str, quote_date: str) -> str:
    text = raw.strip()
    year = quote_date[:4] if re.match(r"\d{4}\.", quote_date) else str(date.today().year)
    match = re.match(r"(\d{1,2})월\s*(\d{1,2})일(.*)", text)
    if not match:
        return text
    return f"{year}.{int(match.group(1)):02}.{int(match.group(2)):02}{match.group(3)}"


def build_order_item_name(lines: list[str]) -> str:
    if not lines:
        return ""

    joined = " ".join(line.strip() for line in lines if line.strip())
    product_line = lines[0].strip()
    model_match = re.search(r"/\s*([A-Za-z0-9-]+)", product_line)
    spec_match = re.search(r"규격:\s*([^\s(]+)", joined)
    cell_match = re.search(r"\*+\s*(\d+\s*Cell)", joined, re.IGNORECASE)
    rack_text = "일체형 랙 구성" if "일체형 랙 구성" in joined else ""
    height_match = re.search(r"\*(\d{3,4})mm", joined)

    if product_line.startswith("축전지") and model_match and spec_match:
        cell_text = f" * {cell_match.group(1).replace(' ', '')}" if cell_match else ""
        rack_suffix = f" + {rack_text}" if rack_text else ""
        height_suffix = f" {height_match.group(1)}mm" if height_match else ""
        return f"축전지 / {spec_match.group(1)}({model_match.group(1)}{cell_text}){rack_suffix}{height_suffix}"

    return joined


def extract_quote_items(lines: list[str]) -> list[dict[str, object]]:
    items: list[dict[str, object]] = []
    buffer: list[str] = []
    in_table = False

    for raw_line in lines:
        line = raw_line.strip()
        if not line:
            continue
        if "Category Product Name" in line:
            in_table = True
            buffer = []
            continue
        if not in_table:
            continue
        if line.startswith("결 재") or line.startswith("납 ") or line.startswith("제 품") or line.startswith("상 호"):
            break
        if "합계(V.A.T" in line or line == "-" or re.fullmatch(r"[-\s]+", line):
            continue

        inline_match = re.match(r"(.+?)\s+(\d+)\s+([\d,]+)\s+([\d,]+)\s*$", line)
        if inline_match:
            desc = inline_match.group(1).strip()
            qty = parse_int(inline_match.group(2))
            unit_price = parse_int(inline_match.group(3))
            total = parse_int(inline_match.group(4))
            if desc:
                item_name = desc
            else:
                item_name = build_order_item_name(buffer)
            if item_name:
                items.append(
                    {
                        "품명규격": item_name,
                        "수량": qty,
                        "단가": unit_price,
                        "금액": total,
                    }
                )
            buffer = []
            continue

        amount_match = re.match(r"(\d+)\s+([\d,]+)\s+([\d,]+)\s*$", line)
        if amount_match:
            item_name = build_order_item_name(buffer)
            if item_name:
                items.append(
                    {
                        "품명규격": item_name,
                        "수량": parse_int(amount_match.group(1)),
                        "단가": parse_int(amount_match.group(2)),
                        "금액": parse_int(amount_match.group(3)),
                    }
                )
            buffer = []
            continue

        if not re.fullmatch(r"[\d,]+", line):
            buffer.append(line)

    return items


def extract_pdf_text(pdf_bytes: bytes) -> tuple[str, list[str]]:
    try:
        from pypdf import PdfReader
    except ImportError as exc:
        raise RuntimeError("PDF 처리 라이브러리(pypdf)가 없습니다. requirements.txt 설치 또는 앱 폴더의 pypdf 포함이 필요합니다.") from exc

    reader = PdfReader(io.BytesIO(pdf_bytes))
    text = "\n".join(page.extract_text() or "" for page in reader.pages)
    lines = [line.strip() for line in text.splitlines() if line.strip()]
    return text, lines


def extract_quote_pdf(pdf_bytes: bytes) -> dict[str, object]:
    text, lines = extract_pdf_text(pdf_bytes)

    quote_no = first_match(r"NO:\s*(MGC-\s*\d+)", text).replace(" ", "")
    date_match = re.search(r"(\d{4})\s*년\s*(\d{1,2})\s*월\s*(\d{1,2})\s*일", text)
    quote_date = format_date_parts(*date_match.groups()) if date_match else ""

    recipient = first_match(r"\n\s*([^\n]+?)\s+귀중", text)
    customer = ""
    contact = ""
    if "/" in recipient:
        customer, contact = [part.strip() for part in recipient.split("/", 1)]
    else:
        customer = recipient.strip()
    contact = contact.replace("님", "").strip()

    site = first_match(r"\n\s*\(([^()\n]+)\)\s*\n\s*아래", text)
    delivery_raw = first_match(r"납\s*기\s*:\s*([^\n]+)", text)
    delivery_place_raw = first_match(r"납\s*품\s*장\s*소\s*:\s*([^\n]+)", text)
    delivery_place = delivery_place_raw
    receiver = ""
    if "담당자:" in delivery_place_raw:
        delivery_place, receiver = [part.strip() for part in delivery_place_raw.split("담당자:", 1)]
        delivery_place = delivery_place.rstrip(",").strip()

    product_config = first_match(r"제\s*품\s*구\s*성\s*:\s*([^\n]+)", text)
    total_amount = parse_int(first_match(r"합계\(V\.A\.T별도\)\s*([\d,]+)", text))
    items = extract_quote_items(lines)
    if "폐전지 회수" in product_config:
        for item in items:
            name = str(item.get("품명규격", "")).strip()
            if name.startswith("설치공사비") and "폐전지" not in name:
                detail = name.replace("설치공사비", "", 1).strip()
                item["품명규격"] = f"설치공사비({detail}+ 폐전지 회수)" if detail else "설치공사비(폐전지 회수)"
    if not total_amount:
        total_amount = sum(parse_int(item.get("금액")) for item in items)

    return {
        "견적번호": quote_no,
        "견적일자": quote_date,
        "발주처": customer,
        "담당자": contact,
        "현장명": site,
        "발행일자": date.today().strftime("%Y.%m.%d"),
        "납품일자": normalize_delivery_date(delivery_raw, quote_date),
        "주문번호": default_order_no(),
        "납품장소": delivery_place,
        "현지인수자": receiver,
        "제품정보": product_config,
        "합계금액": total_amount,
        "품목": items,
    }


def find_purchase_company(text: str) -> str:
    matches = re.findall(r"\n\s*((?:㈜|\(주\)|주식회사)\s*[^\n]+)\s*\n\s*\d{3}-\d{2}-\d{5}", text)
    for value in matches:
        company = re.sub(r"\s+", " ", value).strip()
        if "메가셀" not in company:
            return company
    return ""


def find_purchase_contact(text: str, customer: str) -> str:
    if customer:
        pattern = re.escape(customer) + r".{0,220}?\n\s*([^\n]*(?:팀|부|과|실)[^\n]*(?:부장|팀장|과장|차장|대리|사원)?)\s*\n\s*([\d,-]+[^\n]*)"
        match = re.search(pattern, text, re.S)
        if match:
            return f"{match.group(1).strip()} / {match.group(2).strip()}"
    match = re.search(r"\n\s*([^\n]*(?:정순열|정재헌)[^\n]*)\s*\n\s*([\d,-]+[^\n]*)", text)
    if match:
        return f"{match.group(1).strip()} / {match.group(2).strip()}"
    return ""


def find_purchase_contract_name(text: str, lines: list[str]) -> str:
    inline = first_match(r"계약명\s+([^\n]+)", text)
    if inline and not inline.startswith(("계약일자", "제품명")) and not re.match(r"\d{4}[.-]\d{2}", inline):
        return inline

    for idx, line in enumerate(lines):
        if line.startswith("부가가치세") and idx > 0:
            candidate = lines[idx - 1].strip()
            if candidate and not re.search(r"^\d{4}[.-]\d{2}", candidate):
                return candidate

    for line in lines:
        if any(token in line for token in ["시설공급", "시스템", "공사", "납품", "구매"]) and not line.startswith(("제품명", "특기사항")):
            if not re.search(r"\d{1,3},\d{3}", line):
                return line.strip()
    return ""


def find_purchase_condition(label: str, text: str, lines: list[str]) -> str:
    invalid = {"계약일자(기간)", "계약명", "제품명", "납품장소", "대금지급조건", "주식회사 메가셀"}
    if label == "하자보증기간":
        for idx, line in enumerate(lines):
            if line == label:
                for candidate in reversed(lines[max(0, idx - 4) : idx]):
                    candidate = candidate.strip()
                    if re.fullmatch(r"\d+\s*년", candidate):
                        return candidate
                if idx > 0 and lines[idx - 1].strip() not in invalid:
                    return lines[idx - 1].strip()

    inline = first_match(label + r"\s+([^\n]+)", text)
    if inline and inline not in invalid and not inline.startswith("대금지급조건"):
        return inline
    for idx, line in enumerate(lines):
        if line == label:
            for candidate in lines[idx + 1 : idx + 6]:
                candidate = candidate.strip()
                if candidate and candidate not in invalid and not re.fullmatch(r"\d{3}-\d{2}-\d{5}", candidate):
                    if label in {"납품장소", "입고요청일"} and "협의" not in candidate:
                        continue
                    return candidate
    match = re.search(r"(지엔텔\s+[^\n]+협의)", text)
    return match.group(1).strip() if match else ""


def extract_purchase_order_items(text: str, lines: list[str]) -> list[dict[str, object]]:
    product_name = ""
    for idx, line in enumerate(lines):
        if re.search(r"\s/\s*[A-Za-z0-9-]+", line) and "@" not in line and "mail" not in line.lower():
            product_name = line.strip()
            if idx + 1 < len(lines) and lines[idx + 1].strip() and not lines[idx + 1].startswith("□"):
                product_name = f"{product_name} {lines[idx + 1].strip()}"
            break

    detail_lines: list[str] = []
    collecting_detail = False
    for line in lines:
        stripped = line.strip()
        if stripped.startswith("□"):
            collecting_detail = True
            detail_lines.append(stripped)
            continue
        if collecting_detail:
            if re.search(r"/\s*[A-Za-z0-9-]+", stripped) or stripped in {"합계금액( DC포함 )", "부가가치세(VAT)", "총합계금액(VAT포함)"}:
                collecting_detail = False
                continue
            if stripped and not re.match(r"^\d+\s+\d+\s+", stripped):
                detail_lines.append(stripped)
    product_detail = "\n".join(detail_lines)

    row_match = re.search(r"\n\s*1\s+(\d+)\s+([A-Za-z가-힣]+)\s+([\d,]+)\s+([\d,]+)", text)
    if not row_match:
        return []

    return [
        {
            "품명규격": product_name or "발주 품목",
            "수량": parse_int(row_match.group(1)),
            "단가": parse_int(row_match.group(3)),
            "금액": parse_int(row_match.group(4)),
            "제품설명": product_detail,
        }
    ]


def extract_purchase_order_pdf(pdf_bytes: bytes) -> dict[str, object]:
    text, lines = extract_pdf_text(pdf_bytes)

    order_date_raw = first_match(r"발주일자[^\d]*(\d{4}[-.]\d{2}[-.]\d{2})", text)
    if not order_date_raw:
        order_date_raw = first_match(r"(\d{4}[-.]\d{2}[-.]\d{2})", text)
    order_date = order_date_raw.replace("-", ".")
    customer = find_purchase_company(text)
    contact = find_purchase_contact(text, customer)
    contract_name = find_purchase_contract_name(text, lines)
    delivery_request = find_purchase_condition("입고요청일", text, lines)
    delivery_place = find_purchase_condition("납품장소", text, lines)
    warranty = find_purchase_condition("하자보증기간", text, lines)
    payment = find_purchase_condition("대금지급조건", text, lines)
    items = extract_purchase_order_items(text, lines)

    product_info_parts = []
    if items and items[0].get("제품설명"):
        product_info_parts.append(str(items[0].get("제품설명", "")).strip())
    if contract_name:
        product_info_parts.append(f"계약명: {contract_name}")
    if warranty:
        product_info_parts.append(f"하자보증기간: {warranty}")
    if payment:
        product_info_parts.append(f"대금지급조건: {payment}")

    return {
        "문서유형": "발주서",
        "견적번호": "",
        "견적일자": order_date,
        "발주처": customer,
        "담당자": contact,
        "현장명": contract_name,
        "발행일자": order_date or date.today().strftime("%Y.%m.%d"),
        "납품일자": delivery_request,
        "주문번호": default_order_no(),
        "납품장소": delivery_place,
        "현지인수자": contact,
        "제품정보": "\n".join(part for part in product_info_parts if part),
        "합계금액": sum(parse_int(item.get("금액")) for item in items),
        "품목": items,
    }


def extract_document_pdf(pdf_bytes: bytes) -> dict[str, object]:
    text, _lines = extract_pdf_text(pdf_bytes)
    if "발주서(Purchase Order)" in text or "발주내역" in text:
        return extract_purchase_order_pdf(pdf_bytes)
    data = extract_quote_pdf(pdf_bytes)
    data["문서유형"] = "견적서"
    return data


def safe_filename(text: str) -> str:
    cleaned = re.sub(r'[\\/:*?"<>|]+', " ", text or "")
    cleaned = re.sub(r"\s+", " ", cleaned).strip()
    return cleaned or "주문서"


def build_order_workbook_bytes(data: dict[str, object], items: pd.DataFrame) -> bytes:
    import openpyxl

    template_path = BASE_DIR / "주문서,생산지시서 양식 예시파일.xlsm"
    if not template_path.exists():
        raise FileNotFoundError(f"주문서 양식 파일을 찾을 수 없습니다: {template_path.name}")

    wb = openpyxl.load_workbook(template_path, keep_vba=True)
    ws = wb["주문서"]

    ws["C3"] = str(data.get("발주처", "")).strip()
    ws["C4"] = str(data.get("담당자", "")).strip()
    ws["E3"] = str(data.get("발행일자", "")).strip()
    ws["E4"] = str(data.get("납품일자", "")).strip()
    ws["C5"] = str(data.get("현장명", "")).strip()
    ws["E5"] = str(data.get("주문번호", "")).strip()
    ws["C22"] = str(data.get("제품정보", "")).strip()
    ws["C26"] = str(data.get("납품장소", "")).strip()
    ws["C28"] = str(data.get("현지인수자", "")).strip()

    for row in range(8, 20):
        ws[f"C{row}"] = None
        ws[f"D{row}"] = None
        ws[f"E{row}"] = None

    clean_items = items.copy()
    clean_items = clean_items.dropna(how="all")
    for idx, (_, item) in enumerate(clean_items.head(12).iterrows(), start=8):
        ws[f"C{idx}"] = str(item.get("품명규격", "")).strip()
        ws[f"D{idx}"] = parse_int(item.get("수량", 0))
        ws[f"E{idx}"] = parse_int(item.get("단가", 0))

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def build_current_inventory(history: pd.DataFrame) -> pd.DataFrame:
    history = prepare_numeric(
        clean_display(history),
        ["입고수량", "미출고수량", "출고수량", "잔고수량"],
    )
    if history.empty:
        return history

    key_cols = [
        col
        for col in ["제품분류", "제품명", "유형", "통신카드", "규격"]
        if col in history.columns
    ]
    if not key_cols:
        return history

    numeric_cols = [
        col
        for col in ["입고수량", "미출고수량", "출고수량", "잔고수량"]
        if col in history.columns
    ]
    work = history.copy()
    work[key_cols] = work[key_cols].fillna("")

    current = work.groupby(key_cols, dropna=False)[numeric_cols].sum().reset_index()
    if "잔고수량" not in current.columns and {"입고수량", "출고수량"}.issubset(current.columns):
        current["잔고수량"] = current["입고수량"] - current["출고수량"]

    latest_cols = [col for col in ["일자", "세부내역"] if col in work.columns]
    if latest_cols:
        latest = work.copy()
        latest["_정렬일자"] = pd.to_datetime(latest.get("일자"), errors="coerce")
        latest = latest.sort_values("_정렬일자").groupby(key_cols, dropna=False).tail(1)
        latest = latest[key_cols + latest_cols].rename(
            columns={"일자": "최종일자", "세부내역": "최근세부내역"}
        )
        current = current.merge(latest, on=key_cols, how="left")

    sort_cols = [col for col in ["제품분류", "제품명", "규격"] if col in current.columns]
    if sort_cols:
        current = current.sort_values(sort_cols)
    return current.reset_index(drop=True)


def parse_schedule_date(value: object) -> date | None:
    text = "" if pd.isna(value) else str(value).strip()
    if not text:
        return None

    direct = pd.to_datetime(text, errors="coerce")
    if pd.notna(direct):
        return direct.date()

    match = re.search(r"(20\d{2})[.\-/년\s]+(\d{1,2})(?:[.\-/월\s]+(\d{1,2}))?", text)
    if not match:
        return None

    year = int(match.group(1))
    month = int(match.group(2))
    day = int(match.group(3) or 1)
    try:
        return date(year, month, day)
    except ValueError:
        return None


def build_pending_battery(transactions: pd.DataFrame) -> pd.DataFrame:
    transactions = prepare_numeric(transactions, ["단가", "수량", "금액"])
    if transactions.empty or "구분" not in transactions.columns:
        return transactions

    pending = transactions[
        transactions["구분"].astype(str).str.contains("미출고", na=False)
    ].copy()
    if pending.empty:
        return pending

    pending["달력일자"] = pending["출고일자"].apply(parse_schedule_date)
    pending["년월"] = pending["달력일자"].apply(lambda x: x.strftime("%Y-%m") if x else "")
    return pending.reset_index(drop=True)


def render_pending_calendar(pending: pd.DataFrame, month_key: str) -> None:
    if pending.empty or not month_key:
        st.info("달력에 표시할 미출고 일정이 없습니다.")
        return

    year, month = [int(part) for part in month_key.split("-")]
    calendar.setfirstweekday(calendar.SUNDAY)

    headers = ["일", "월", "화", "수", "목", "금", "토"]
    for col, label in zip(st.columns(7), headers):
        col.markdown(f"**{label}**")

    scheduled = pending[pending["년월"] == month_key].copy()
    by_day = {day: group for day, group in scheduled.groupby(scheduled["달력일자"].apply(lambda x: x.day))}

    for week in calendar.monthcalendar(year, month):
        cols = st.columns(7)
        for col, day_num in zip(cols, week):
            with col.container(border=True):
                if day_num == 0:
                    st.caption(" ")
                    st.write("")
                    continue
                st.caption(f"{day_num}일")
                day_rows = by_day.get(day_num)
                if day_rows is None or day_rows.empty:
                    st.write("")
                    continue
                for _, row in day_rows.iterrows():
                    qty = normalize_numeric(pd.Series([row.get("수량", 0)])).iloc[0]
                    label = f"{row.get('품목명', '')} / {qty:,.0f}개"
                    memo = str(row.get("메모", "")).strip()
                    order_no = str(row.get("주문서 번호", "")).strip()
                    st.markdown(f"**{label}**")
                    if memo:
                        st.caption(memo)
                    if order_no:
                        st.caption(order_no)


def searchable_table(
    table_name: str,
    title: str,
    search_label: str,
    search_cols: Iterable[str] | None,
    preferred_cols: Iterable[str] | None = None,
    csv_name: str | None = None,
) -> pd.DataFrame:
    st.subheader(title)
    df = clean_display(get_table(table_name))
    keyword = st.text_input(search_label, key=f"search_{table_name}_{title}")
    filtered = text_filter(df, keyword, search_cols)
    st.metric("검색 결과", f"{len(filtered):,}건")
    table_view(filtered, preferred_cols)
    if csv_name:
        download_csv(filtered, csv_name)
    return filtered


PAGE_GROUPS = {
    "대시보드": ["대시보드"],
    "영업관리": ["수주 현황"],
    "경영관리": ["경영 요약"],
    "구매재고관리": ["재고 관리", "재고 출납기록", "자재/BOM", "자재 발주"],
    "배터리관리": ["배터리 관리"],
    "AS관리": ["AS 관리"],
    "문서출력": ["견적/주문서"],
}


def run_db_sync() -> subprocess.CompletedProcess[str]:
    return subprocess.run(
        [sys.executable, str(BASE_DIR / "init_db.py")],
        cwd=BASE_DIR,
        capture_output=True,
        text=True,
        encoding="utf-8",
        errors="replace",
        timeout=300,
    )


if not DB_PATH.exists():
    st.error(f"DB 파일을 찾을 수 없습니다: {DB_PATH}")
    st.info("먼저 같은 폴더에서 `python init_db.py`를 실행해 주세요.")
    st.stop()

st.sidebar.title("메가셀 ERP")
st.sidebar.caption("엑셀 원장 기반 로컬 통합 조회")
st.sidebar.divider()

with st.sidebar.expander("원장/DB 관리", expanded=False):
    st.caption(str(BASE_DIR))
    if st.button("원장 동기화", width="stretch"):
        with st.spinner("엑셀 원장을 DB로 다시 읽는 중입니다..."):
            try:
                result = run_db_sync()
            except subprocess.TimeoutExpired:
                st.error("동기화 시간이 너무 오래 걸려 중단됐습니다.")
            except Exception as exc:
                st.error(f"동기화 실행 오류: {exc}")
            else:
                if result.returncode == 0:
                    st.cache_data.clear()
                    st.success("원장 동기화 완료")
                    with st.expander("동기화 로그"):
                        st.code(result.stdout[-4000:] or "완료")
                    st.rerun()
                else:
                    st.error("원장 동기화 실패")
                    st.code((result.stdout + "\n" + result.stderr)[-5000:])
    if st.button("화면 새로고침", width="stretch"):
        st.cache_data.clear()
        st.rerun()

st.sidebar.divider()
section = st.sidebar.radio("업무영역", list(PAGE_GROUPS.keys()))
pages = PAGE_GROUPS[section]
menu = pages[0] if len(pages) == 1 else st.sidebar.radio("화면", pages)

available_tables = list_tables(db_mtime())
required_tables = ["수주", "재고", "재고출납", "AS관리", "부품리스트", "자재재고", "BOM", "발주현황"]
missing_tables = [name for name in required_tables if name not in available_tables]
if missing_tables:
    st.warning("DB에 아직 없는 테이블: " + ", ".join(missing_tables))


if menu == "대시보드":
    st.title("대시보드")
    st.caption("현재 DB에 적재된 업무 원장 기준 요약입니다.")

    orders = prepare_numeric(get_table("수주"), ["수량", "단가", "수주금액", "vat포함"])
    inventory_stock = prepare_numeric(get_table("재고"), ["입고수량", "출고수량", "미출고수량", "현재고"])
    as_history = prepare_numeric(get_table("AS관리"), ["AS비용"])
    material_stock = prepare_numeric(get_table("자재재고"), ["입고량", "출고량", "현재고", "재고금액"])
    purchases = prepare_numeric(get_table("발주현황"), ["개수", "단가", "총금액"])

    col1, col2, col3, col4, col5 = st.columns(5)
    col1.metric("수주", f"{len(orders):,}건")
    col2.metric("수주금액", metric_sum(orders, "수주금액", "원"))
    col3.metric("제품 현재고", metric_sum(inventory_stock, "현재고"))
    col4.metric("자재 현재고", metric_sum(material_stock, "현재고"))
    col5.metric("AS 이력", f"{len(as_history):,}건")

    st.divider()
    left, right = st.columns(2)
    with left:
        st.subheader("최근 수주")
        table_view(
            orders.tail(12).iloc[::-1],
            ["수주일자", "납품처", "품명규격", "수량", "수주금액", "출고일", "계산서"],
            height=360,
        )
    with right:
        st.subheader("최근 자재 발주")
        table_view(
            purchases.tail(12).iloc[::-1],
            ["업체명", "발주일", "입고(요청)일", "품목", "품명", "개수", "진행 상태"],
            height=360,
        )


elif menu == "수주 현황":
    st.title("수주 현황")
    orders = prepare_numeric(get_table("수주"), ["수량", "단가", "수주금액", "vat포함"])

    c1, c2, c3, c4 = st.columns([2, 2, 1.5, 1.5])
    customer_kw = c1.text_input("납품처/발주처 검색")
    product_kw = c2.text_input("품명/규격 검색")
    instruction_kw = c3.text_input("제조지시서번호 검색")
    order_date_kw = c4.text_input("수주일자 검색")
    c5, c6 = st.columns([1.4, 1.4])
    category = c5.selectbox("구분", ["전체"] + safe_unique(orders, "구분"))
    invoice = c6.selectbox("계산서", ["전체"] + safe_unique(orders, "계산서"))

    filtered = text_filter(orders, customer_kw, ["납품처", "발주처"])
    filtered = text_filter(filtered, product_kw, ["품명규격", "비고"])
    filtered = text_filter(filtered, instruction_kw, ["제조지시서"])
    filtered = text_filter(filtered, order_date_kw, ["수주일자"])
    if category != "전체" and "구분" in filtered.columns:
        filtered = filtered[filtered["구분"].astype(str) == category]
    if invoice != "전체" and "계산서" in filtered.columns:
        filtered = filtered[filtered["계산서"].astype(str) == invoice]

    m1, m2, m3, m4 = st.columns(4)
    m1.metric("검색 결과", f"{len(filtered):,}건")
    m2.metric("수량 합계", metric_sum(filtered, "수량"))
    m3.metric("수주금액 합계", metric_sum(filtered, "수주금액", "원"))
    m4.metric("VAT 포함 합계", metric_sum(filtered, "vat포함", "원"))

    st.divider()
    table_view(
        filtered,
        [
            "월",
            "구분",
            "수주일자",
            "제조지시서",
            "납품처",
            "품명규격",
            "단위",
            "수량",
            "단가",
            "수주금액",
            "vat포함",
            "발주처",
            "출고요청일",
            "출고일",
            "계산서",
            "비고",
        ],
    )
    download_csv(filtered, "megacell_orders.csv")


elif menu == "경영 요약":
    st.title("경영 요약")
    st.caption("현재 적재된 업무 원장 기준의 금액·건수 요약입니다.")

    orders = prepare_numeric(get_table("수주"), ["수량", "수주금액", "vat포함"])
    purchases = prepare_numeric(get_table("발주현황"), ["개수", "단가", "총금액"])
    material_stock = prepare_numeric(get_table("자재재고"), ["현재고", "재고금액"])
    as_history = prepare_numeric(get_table("AS관리"), ["AS비용"])

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("수주금액", metric_sum(orders, "수주금액", "원"))
    c2.metric("VAT 포함", metric_sum(orders, "vat포함", "원"))
    c3.metric("자재 발주금액", metric_sum(purchases, "총금액", "원"))
    c4.metric("AS 비용", metric_sum(as_history, "AS비용", "원"))

    st.divider()
    left, right = st.columns(2)

    with left:
        st.subheader("월별 수주 요약")
        if orders.empty or "월" not in orders.columns:
            st.info("수주 월 데이터가 없습니다.")
        else:
            monthly = (
                orders.assign(월=orders["월"].astype(str))
                .groupby("월", dropna=False)
                .agg(건수=("월", "size"), 수주금액=("수주금액", "sum"), vat포함=("vat포함", "sum"))
                .reset_index()
                .sort_values("월")
            )
            table_view(monthly, ["월", "건수", "수주금액", "vat포함"], height=360)

    with right:
        st.subheader("거래처별 수주 요약")
        if orders.empty or "납품처" not in orders.columns:
            st.info("납품처 데이터가 없습니다.")
        else:
            by_customer = (
                orders.groupby("납품처", dropna=False)
                .agg(건수=("납품처", "size"), 수주금액=("수주금액", "sum"), vat포함=("vat포함", "sum"))
                .reset_index()
                .sort_values("수주금액", ascending=False)
                .head(30)
            )
            table_view(by_customer, ["납품처", "건수", "수주금액", "vat포함"], height=360)

    st.divider()
    st.subheader("구매·재고 금액 요약")
    summary = pd.DataFrame(
        [
            {"구분": "자재 발주", "건수": len(purchases), "금액": normalize_numeric(purchases.get("총금액", pd.Series(dtype=float))).sum()},
            {"구분": "자재 재고", "건수": len(material_stock), "금액": normalize_numeric(material_stock.get("재고금액", pd.Series(dtype=float))).sum()},
            {"구분": "AS 비용", "건수": len(as_history), "금액": normalize_numeric(as_history.get("AS비용", pd.Series(dtype=float))).sum()},
        ]
    )
    table_view(summary, ["구분", "건수", "금액"], height=220)
    download_csv(summary, "megacell_management_summary.csv")


elif menu == "재고 관리":
    st.title("재고 관리")
    inventory = prepare_numeric(get_table("재고"), ["입고수량", "미출고수량", "출고수량", "현재고"])

    c1, c2, c3 = st.columns([2, 1.4, 1.4])
    keyword = c1.text_input("제품명/규격 검색")
    product_type = c2.selectbox("제품분류", ["전체"] + safe_unique(inventory, "제품분류"))
    stock_filter = c3.selectbox("재고 상태", ["전체", "재고 있음", "재고 없음"])

    filtered = text_filter(inventory, keyword, ["제품명", "규격", "유형", "통신카드", "비고"])
    if product_type != "전체" and "제품분류" in filtered.columns:
        filtered = filtered[filtered["제품분류"].astype(str) == product_type]
    if stock_filter == "재고 있음" and "현재고" in filtered.columns:
        filtered = filtered[filtered["현재고"] > 0]
    if stock_filter == "재고 없음" and "현재고" in filtered.columns:
        filtered = filtered[filtered["현재고"] <= 0]

    m1, m2, m3, m4 = st.columns(4)
    m1.metric("품목 수", f"{len(filtered):,}개")
    m2.metric("입고수량", metric_sum(filtered, "입고수량"))
    m3.metric("출고수량", metric_sum(filtered, "출고수량"))
    m4.metric("현재고", metric_sum(filtered, "현재고"))

    st.divider()
    table_view(
        filtered,
        [
            "제품분류",
            "제품명",
            "유형",
            "통신카드",
            "규격",
            "입고수량",
            "미출고수량",
            "출고수량",
            "현재고",
            "비고",
        ],
    )
    download_csv(filtered, "megacell_current_inventory.csv")


elif menu == "재고 출납기록":
    st.title("재고 출납기록")
    inventory = prepare_numeric(get_table("재고출납"), ["입고수량", "출고수량", "미출고수량", "잔고수량"])
    if inventory.empty and "재고출납" not in available_tables:
        st.info("재고 출납기록 테이블이 아직 없습니다. 원장 동기화를 한 번 실행해 주세요.")

    c1, c2, c3 = st.columns([2, 1.4, 1.4])
    keyword = c1.text_input("제품명/규격/세부내역 검색")
    product_type = c2.selectbox("제품분류", ["전체"] + safe_unique(inventory, "제품분류"))
    work_type = c3.selectbox("업무구분", ["전체"] + safe_unique(inventory, "업무구분"))

    filtered = text_filter(inventory, keyword, ["제품명", "규격", "세부내역", "비고"])
    if product_type != "전체" and "제품분류" in filtered.columns:
        filtered = filtered[filtered["제품분류"].astype(str) == product_type]
    if work_type != "전체" and "업무구분" in filtered.columns:
        filtered = filtered[filtered["업무구분"].astype(str) == work_type]

    m1, m2, m3, m4 = st.columns(4)
    m1.metric("검색 결과", f"{len(filtered):,}건")
    m2.metric("입고수량", metric_sum(filtered, "입고수량"))
    m3.metric("출고수량", metric_sum(filtered, "출고수량"))
    m4.metric("잔고수량", metric_sum(filtered, "잔고수량"))

    st.divider()
    table_view(
        filtered,
        [
            "재고정보코드",
            "일자",
            "제품분류",
            "제품명",
            "유형",
            "통신카드",
            "규격",
            "업무구분",
            "세부내역",
            "입고수량",
            "미출고수량",
            "출고수량",
            "잔고수량",
            "비고",
        ],
    )
    download_csv(filtered, "megacell_inventory_history.csv")


elif menu == "자재/BOM":
    st.title("자재/BOM")
    tab_stock, tab_parts, tab_bom = st.tabs(["자재 재고", "부품리스트", "BOM"])

    with tab_stock:
        stock = prepare_numeric(get_table("자재재고"), ["단가", "입고량", "출고량", "현재고", "재고금액"])
        keyword = st.text_input("제조사/자재품목/자재품명 검색", key="material_stock_search")
        filtered = text_filter(stock, keyword, ["제조사", "자재품목", "자재품명"])
        m1, m2, m3 = st.columns(3)
        m1.metric("검색 결과", f"{len(filtered):,}건")
        m2.metric("현재고 합계", metric_sum(filtered, "현재고"))
        m3.metric("재고금액", metric_sum(filtered, "재고금액", "원"))
        table_view(filtered, ["제조사", "자재품목", "자재품명", "단가", "단위", "입고량", "출고량", "현재고", "재고금액"])
        download_csv(filtered, "megacell_material_stock.csv")

    with tab_parts:
        searchable_table(
            "부품리스트",
            "부품리스트",
            "자재품목/자재품명/제조사 검색",
            ["자재품목", "자재품명", "제조사", "자재용도"],
            ["자재품목", "자재품명", "제조사", "단위", "단가", "자재용도"],
            "megacell_parts.csv",
        )

    with tab_bom:
        st.subheader("BOM 관리")
        bom = prepare_numeric(get_table("BOM"), ["소요량", "단가"])

        if bom.empty:
            st.info("BOM 데이터가 없습니다.")
        else:
            c1, c2, c3 = st.columns([1.4, 1.8, 1])
            category_options = safe_unique(bom, "제품 분류")
            selected_category = c1.selectbox("제품분류", category_options, key="bom_category")

            category_bom = bom[bom["제품 분류"].astype(str) == selected_category] if selected_category else bom
            product_options = safe_unique(category_bom, "완제품명")
            selected_product = c2.selectbox("완제품명", product_options, key="bom_product")
            production_qty = c3.number_input("대수", min_value=1, value=1, step=1, key="bom_qty")

            selected_bom = category_bom[
                category_bom["완제품명"].astype(str) == selected_product
            ].copy()

            result = prepare_numeric(selected_bom.copy(), ["소요량", "단가"])
            result["대수"] = production_qty
            if "소요량" in result.columns:
                result["필요수량"] = result["소요량"] * production_qty
            if {"필요수량", "단가"}.issubset(result.columns):
                result["예상금액"] = result["필요수량"] * result["단가"]

            original_col, result_col = st.columns(2, gap="large")
            original_cols = [
                "제품 분류",
                "완제품명",
                "자재품목",
                "자재품명",
                "제조사",
                "단위",
                "소요량",
                "단가",
                "자재용도",
            ]
            result_cols = [
                "자재품목",
                "자재품명",
                "제조사",
                "단위",
                "소요량",
                "대수",
                "필요수량",
                "단가",
                "예상금액",
                "자재용도",
            ]

            with original_col:
                st.markdown("#### BOM 원본")
                with st.container(border=True):
                    st.caption("엑셀 BOM 원장에서 읽어온 1대 기준값입니다.")
                    table_view(selected_bom, original_cols, height=620)

            with result_col:
                st.markdown("#### 대수 반영 결과")
                with st.container(border=True):
                    st.caption(f"{production_qty:,}대 기준으로 계산한 필요수량과 예상금액입니다.")
                    m1, m2, m3 = st.columns(3)
                    m1.metric("자재 항목", f"{len(result):,}개")
                    m2.metric("총 필요수량", metric_sum(result, "필요수량"))
                    m3.metric("예상금액", metric_sum(result, "예상금액", "원"))
                    table_view(result, result_cols, height=520)
                    download_csv(result, "megacell_bom_adjusted.csv")


elif menu == "자재 발주":
    st.title("자재 발주")
    purchases = prepare_numeric(get_table("발주현황"), ["개수", "단가", "총금액"])

    c1, c2 = st.columns([2, 1.4])
    keyword = c1.text_input("업체명/품목/품명/발주목적 검색")
    status = c2.selectbox("진행 상태", ["전체"] + safe_unique(purchases, "진행 상태"))
    filtered = text_filter(purchases, keyword, ["업체명", "품목", "품명", "발주목적", "비고"])
    if status != "전체" and "진행 상태" in filtered.columns:
        filtered = filtered[filtered["진행 상태"].astype(str) == status]

    m1, m2, m3 = st.columns(3)
    m1.metric("검색 결과", f"{len(filtered):,}건")
    m2.metric("발주 수량", metric_sum(filtered, "개수"))
    m3.metric("총금액", metric_sum(filtered, "총금액", "원"))
    st.divider()
    table_view(
        filtered,
        ["No.", "업체명", "발주일", "입고(요청)일", "발주번호", "품목", "품명", "개수", "단가", "총금액", "발주목적", "진행 상태", "비고"],
    )
    download_csv(filtered, "megacell_purchases.csv")


elif menu == "배터리 관리":
    st.title("배터리 관리")
    stock = prepare_numeric(get_table("배터리재고"), ["입고", "출고", "미출고", "재고"])
    transactions = prepare_numeric(get_table("배터리입출고"), ["단가", "수량", "금액"])
    items = prepare_numeric(get_table("배터리품목"), ["단가"])
    usage_log = get_table("배터리사용이력")

    if stock.empty and transactions.empty and items.empty and usage_log.empty:
        st.info("배터리 데이터가 아직 DB에 적재되지 않았습니다. `python init_db.py`를 다시 실행해 주세요.")
        st.stop()

    tab_stock, tab_history, tab_schedule, tab_items, tab_log = st.tabs(
        ["현재고", "입출고 이력", "미출고 달력", "품목관리", "사용이력"]
    )

    with tab_stock:
        keyword = st.text_input("품목명 검색", key="battery_stock_search")
        filtered = text_filter(stock, keyword, ["품목명"])
        c1, c2, c3, c4 = st.columns(4)
        c1.metric("품목 수", f"{len(filtered):,}개")
        c2.metric("입고", metric_sum(filtered, "입고"))
        c3.metric("출고", metric_sum(filtered, "출고"))
        c4.metric("재고", metric_sum(filtered, "재고"))
        table_view(filtered, ["품목명", "입고", "출고", "미출고", "재고"])
        download_csv(filtered, "megacell_battery_stock.csv")

    with tab_history:
        c1, c2 = st.columns([2, 1.3])
        keyword = c1.text_input("품목명/메모/주문서 번호 검색", key="battery_history_search")
        kind = c2.selectbox("구분", ["전체"] + safe_unique(transactions, "구분"))
        filtered = text_filter(transactions, keyword, ["품목명", "메모", "주문서 번호", "품목코드"])
        if kind != "전체" and "구분" in filtered.columns:
            filtered = filtered[filtered["구분"].astype(str) == kind]
        m1, m2, m3 = st.columns(3)
        m1.metric("검색 결과", f"{len(filtered):,}건")
        m2.metric("수량 합계", metric_sum(filtered, "수량"))
        m3.metric("금액 합계", metric_sum(filtered, "금액", "원"))
        table_view(
            filtered,
            ["ID", "작성일자", "구분", "출고일자", "품목코드", "품목명", "단가", "수량", "금액", "메모", "주문서 번호"],
        )
        download_csv(filtered, "megacell_battery_transactions.csv")

    with tab_schedule:
        pending = build_pending_battery(transactions)
        parsed_months = sorted([value for value in pending.get("년월", pd.Series(dtype=str)).unique() if value])
        default_month = date.today().strftime("%Y-%m")
        default_index = parsed_months.index(default_month) if default_month in parsed_months else 0

        c1, c2 = st.columns([1.2, 2])
        month_key = c1.selectbox(
            "월",
            parsed_months,
            index=default_index if parsed_months else None,
            key="battery_pending_month",
        )
        keyword = c2.text_input("품목명/메모/주문서 번호 검색", key="battery_pending_search")

        filtered = text_filter(pending, keyword, ["품목명", "메모", "주문서 번호", "품목코드"])
        m1, m2, m3 = st.columns(3)
        m1.metric("미출고", f"{len(filtered):,}건")
        m2.metric("수량 합계", metric_sum(filtered, "수량"))
        m3.metric("금액 합계", metric_sum(filtered, "금액", "원"))

        st.divider()
        render_pending_calendar(filtered, month_key)

        unscheduled = filtered[filtered["달력일자"].isna()].copy()
        if not unscheduled.empty:
            st.divider()
            st.subheader("일정 미확정")
            table_view(
                unscheduled,
                ["ID", "작성일자", "출고일자", "품목코드", "품목명", "수량", "금액", "메모", "주문서 번호"],
                height=260,
            )

        st.divider()
        st.subheader("미출고 목록")
        table_view(
            filtered,
            ["ID", "작성일자", "출고일자", "품목코드", "품목명", "수량", "금액", "메모", "주문서 번호"],
            height=360,
        )
        download_csv(filtered, "megacell_battery_pending.csv")

    with tab_items:
        keyword = st.text_input("품목코드/품목명 검색", key="battery_items_search")
        filtered = text_filter(items, keyword, ["품목코드", "품목명"])
        st.metric("품목 수", f"{len(filtered):,}개")
        table_view(filtered, ["품목코드", "품목명", "단가"])
        download_csv(filtered, "megacell_battery_items.csv")

    with tab_log:
        keyword = st.text_input("작업 내역 검색", key="battery_log_search")
        filtered = text_filter(usage_log, keyword, ["구분", "작업 내역 상세"])
        st.metric("사용이력", f"{len(filtered):,}건")
        table_view(filtered, ["일시", "구분", "작업 내역 상세"])
        download_csv(filtered, "megacell_battery_usage_log.csv")


elif menu == "AS 관리":
    st.title("AS 관리")
    as_history = prepare_numeric(get_table("AS관리"), ["AS비용"])

    c1, c2 = st.columns([2, 1.4])
    keyword = c1.text_input("접수처/제품명/SN/증상 검색")
    status = c2.selectbox("상태", ["전체"] + safe_unique(as_history, "상태"))
    filtered = text_filter(as_history, keyword, ["접수처", "제품명", "SN", "증상내용", "관리번호"])
    if status != "전체" and "상태" in filtered.columns:
        filtered = filtered[filtered["상태"].astype(str) == status]

    m1, m2, m3 = st.columns(3)
    m1.metric("검색 결과", f"{len(filtered):,}건")
    m2.metric("전체 AS 이력", f"{len(as_history):,}건")
    m3.metric("AS 비용 합계", metric_sum(filtered, "AS비용", "원"))

    st.divider()
    table_view(
        filtered,
        [
            "관리번호",
            "접수일",
            "특수관리",
            "접수처",
            "제품명",
            "SN",
            "수량",
            "증상내용",
            "처리방법",
            "유무상",
            "AS비용",
            "비용처리상태",
            "상태",
            "담당자",
            "출고일",
            "비고",
        ],
    )
    download_csv(filtered, "megacell_as.csv")


elif menu == "견적/주문서":
    st.title("견적/주문서")
    tab_pdf, tab_orders, tab_templates = st.tabs(["문서 PDF 변환", "수주 원장 조회", "양식 파일"])

    with tab_pdf:
        st.subheader("견적서/발주서 PDF -> 주문서")
        uploaded_pdf = st.file_uploader("견적서 또는 발주서 PDF 업로드", type=["pdf"], key="quote_pdf_upload")

        if uploaded_pdf is None:
            st.info("텍스트형 견적서 또는 발주서 PDF를 업로드하면 주문서 양식에 들어갈 값을 추출합니다.")
        else:
            try:
                extracted = extract_document_pdf(uploaded_pdf.getvalue())
            except Exception as exc:
                st.error(f"PDF 추출 실패: {exc}")
            else:
                st.caption(
                    f"문서유형: {extracted.get('문서유형') or '-'} / "
                    f"문서번호: {extracted.get('견적번호') or '-'} / "
                    f"문서일자: {extracted.get('견적일자') or '-'}"
                )

                c1, c2, c3 = st.columns([1.4, 1.2, 1.2])
                extracted["발주처"] = c1.text_input("발주처", value=str(extracted.get("발주처", "")))
                extracted["담당자"] = c2.text_input("담당자", value=str(extracted.get("담당자", "")))
                extracted["현장명"] = c3.text_input("P/O No. 또는 현장명", value=str(extracted.get("현장명", "")))

                c4, c5, c6 = st.columns([1.2, 1.8, 1.2])
                extracted["발행일자"] = c4.text_input("주문서 발행일자", value=str(extracted.get("발행일자", "")))
                extracted["납품일자"] = c5.text_input("납품일자/시간", value=str(extracted.get("납품일자", "")))
                extracted["주문번호"] = c6.text_input("S/O No.", value=str(extracted.get("주문번호", "")))

                extracted["납품장소"] = st.text_input("납품 또는 설치 장소", value=str(extracted.get("납품장소", "")))
                extracted["현지인수자"] = st.text_input("현지 인수자", value=str(extracted.get("현지인수자", "")))
                extracted["제품정보"] = st.text_area("제품정보 및 첨부품", value=str(extracted.get("제품정보", "")), height=90)

                st.subheader("품목 확인")
                items_df = pd.DataFrame(extracted.get("품목") or [])
                if items_df.empty:
                    items_df = pd.DataFrame(columns=["품명규격", "수량", "단가", "금액"])
                edited_items = st.data_editor(
                    items_df,
                    width="stretch",
                    hide_index=True,
                    num_rows="dynamic",
                    column_order=["품명규격", "수량", "단가", "금액"],
                    key="quote_items_editor",
                )
                edited_items = prepare_numeric(edited_items, ["수량", "단가", "금액"])

                m1, m2 = st.columns(2)
                m1.metric("품목 수", f"{len(edited_items.dropna(how='all')):,}개")
                m2.metric("품목 합계", metric_sum(edited_items, "금액", "원"))

                try:
                    order_bytes = build_order_workbook_bytes(extracted, edited_items)
                except Exception as exc:
                    st.error(f"주문서 생성 실패: {exc}")
                else:
                    file_base = safe_filename(
                        f"{extracted.get('주문번호', '')} {extracted.get('발주처', '')} "
                        f"{extracted.get('담당자', '')} {extracted.get('현장명', '')}"
                    )
                    st.download_button(
                        "주문서 xlsm 내려받기",
                        data=order_bytes,
                        file_name=f"{file_base}.xlsm",
                        mime="application/vnd.ms-excel.sheet.macroEnabled.12",
                        width="stretch",
                    )

    with tab_orders:
        orders = prepare_numeric(get_table("수주"), ["수량", "수주금액", "vat포함"])
        keyword = st.text_input("문서 생성 대상 검색")
        filtered = text_filter(orders, keyword, ["납품처", "품명규격", "제조지시서", "발주처"])
        table_view(
            filtered.head(100),
            ["수주일자", "제조지시서", "납품처", "품명규격", "수량", "수주금액", "vat포함", "발주처", "비고"],
            height=500,
        )

    with tab_templates:
        st.subheader("포함된 양식 파일")
        templates = sorted(
            [path.name for path in BASE_DIR.iterdir() if path.suffix.lower() in {".pdf", ".xlsm"} and "양식" in path.name]
        )
        template_df = pd.DataFrame({"파일명": templates})
        table_view(template_df, ["파일명"], height=220)
